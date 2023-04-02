# Python program to read an excel file

import datetime
# import openpyxl module
import openpyxl
 
# Give the location of the file
path = "./ADAT.xlsx"
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
def retorno_de_spreadsheet(sheet_obj):
    # Cell object is created by using
    # sheet object's cell() method.
    cell_obj = sheet_obj.cell(row = 1, column = 1)
    
    # Print value of cell object
    # using the value attribute
    lista = [[]]
    print(sheet_obj.max_column)
    max_col = sheet_obj.max_column
    m_row = sheet_obj.max_row
    # Loop will print all columns name
    for i in range(2, m_row + 1):
        for j in range(1, max_col + 1):    
            cell_obj = sheet_obj.cell(row = i, column = j)
            lista[i-2].append(cell_obj.value)
        lista.append([])
    return [[item if not None else "" for item in linha] for linha in lista if any(elem is not None for elem in linha)]
membro = retorno_de_spreadsheet(wb_obj['MEMBROS'])
endereco = retorno_de_spreadsheet(wb_obj['ENDERECO'])
tel = retorno_de_spreadsheet(wb_obj['TELEFONE'])

#-----------------------------------------------

import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database = "ADAT",
  auth_plugin='mysql_native_password'
)
cursor = mydb.cursor()
for index, members in enumerate(membro):
    data = (members[0], members[1], members[2], members[3],members[4].isoformat().split('T')[0])
    # INSERT INTO MEMBROS VALUES(NULL,'Rafael Palheta de Lima', '010.596-252-02', 'O+','rpalhetavigia@gmail.com', '26/03/22');
    add_member = ("INSERT INTO MEMBROS "
                "VALUES (NULL, %s, %s, %s, %s, %s)") 
    cursor.execute(add_member, data)
    member_no = cursor.lastrowid
    add_end = ("INSERT INTO ENDERECO "
            "VALUES (NULL, %s, %s, %s, %s, %s)")
    add_tel = ("INSERT INTO TELEFONE "
            "VALUES (NULL, %s, %s, %s, %s)")
    data_end = (endereco[index][0], endereco[index][1],endereco[index][2],endereco[index][3], member_no)
    cursor.execute(add_end, data_end)
    data_tel = (tel[index][0], tel[index][1],tel[index][2], member_no) 
    cursor.execute(add_tel, data_tel)
    mydb.commit()
cursor.close()
mydb.close()